import requests
import xml.etree.ElementTree as ET
from logger import logger
import openpyxl
import xlrd
import os

class ProductsCGU:
    def __init__(self):
        super().__init__()
        self.log = logger()

    def getProductsCGU(self, url, username, password):
        try:
            headers = {
                "Content-Type": "text/html",
                "SOAPAction": "http://nodum.com.uy/soap/schemas/forms/v1.2/WSArticulosSOFT/resp/procesarAlta"
            }

            body = """
            <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:resp="http://nodum.com.uy/soap/schemas/forms/v1.2/WSArticulosSOFT/resp">
            <soapenv:Header/>
            <soapenv:Body>
                <resp:procesarAlta>
                    <!--1 or more repetitions:-->
                    <resp:WSArticulosSOFT/>
                </resp:procesarAlta>
            </soapenv:Body>
            </soapenv:Envelope>
            """

            response = requests.post(url, data=body, headers=headers, auth=(username, password))
            # print(f"productos.py - Ln40: {response.status_code}")
            if response.status_code == 200:
                root = ET.fromstring(response.text)
                
                namespace = {
                    'soapenv': 'http://schemas.xmlsoap.org/soap/envelope/',
                    'ns1': 'http://nodum.com.uy/soap/schemas/forms/v1.2/WSArticulosSOFT/resp'
                }
                
                general_data = root.findall('.//ns1:General/ns1:Articulos', namespace)
                
                general_list = []
                for general in general_data:
                    codigo = general.find('ns1:cod_articulo', namespace).text
                    nombre = general.find('ns1:nom_articulo', namespace).text
                    precio = general.find('ns1:precio', namespace).text
                    tipo_articulo = general.find('ns1:tipo_articulo', namespace).text
                    grupo = general.find('ns1:NombreFamiliaArticulo', namespace).text

                    if tipo_articulo == 'Articulo': #Tengo que ver que este activo
                        precio_sin_iva = round(float(precio)) / float(f"1.22")
                        general_list.append({
                            "idproducto": codigo,
                            "descripcion": nombre.upper(),
                            "plu" : "",
                            "impuesto" : "",
                            "iva" : "22",
                            "unidad" : "ACTIVIDAD",
                            "precio_sin_iva": precio_sin_iva,
                            "precio_iva_inc": round(float(precio)),
                            "importe_iva": round(round(float(precio)) - precio_sin_iva, 2),
                            "grupo" : grupo.upper()
                        })

                return self.setDiccionary(general_list)
        
            else:
                self.log.write_to_log(f"====================================\nError ClientsCGU: {response.status_code}\n {response.text} \n====================================")
        except Exception as e:
            self.log.write_to_log(f"(getProductsCGU) - Se produjo un error: {str(e)}")
    
    def setDiccionary(self, cguproducts):
        grupos_dict = {}
        for producto in cguproducts:
            grupo = producto['grupo']
            idproducto = producto['idproducto']

            if grupo not in grupos_dict:
                grupos_dict[grupo] = {
                    "grupo": grupo,
                    "productos": {}
                }

            grupos_dict[grupo]["productos"][idproducto] = {
                "idproducto": idproducto,
                "descripcion": producto['descripcion'],
                "plu": producto['plu'],
                "impuesto": producto['impuesto'],
                "iva": producto['iva'],
                "unidad": producto['unidad'],
                "precio_sin_iva": producto['precio_sin_iva'],
                "precio_iva_inc": producto['precio_iva_inc'],
                "importe_iva": producto['importe_iva'],
            }

        return grupos_dict

class ClientsCGU:

    def __init__(self):
        super().__init__()
        self.log = logger()

    def getClientsCGU(self, file_name):
        try:
            grupos_dict = []
            file_extension = os.path.splitext(file_name)[1].lower()

            if file_extension == '.xls':
                workbook = xlrd.open_workbook(file_name)
                worksheet = workbook.sheet_by_index(0)
                for row_idx in range(1, worksheet.nrows):
                    self._process_row(grupos_dict, worksheet.row_values(row_idx))
            elif file_extension == '.xlsx':
                workbook = openpyxl.load_workbook(file_name, read_only=True, data_only=True)
                worksheet = workbook[workbook.sheetnames[0]]
                for row in worksheet.iter_rows(min_row=3, values_only=True):
                    self._process_row(grupos_dict, row)
            else:
                raise ValueError("Formato de archivo no soportado. Usa .xls o .xlsx")
        
            return grupos_dict
        except Exception as e:
            self.log.write_to_log(f"(getClientsCGU) - Se produjo un error: {str(e)}")
            return []  # Retorna lista vacía en caso de error

    def _process_row(self, grupos_dict, row):
        if len(row) < 2:
            self.log.write_to_log(f"(getClientsCGU) - Fila con datos insuficientes: {row}")
            return
        
        cedula = str(row[0]).strip()  # Convertir a string y limpiar espacios
        nombre = str(row[1]).strip()

        grupos_dict.append({
            "cedula": cedula.upper(),
            "nombre": nombre.upper(),
        })